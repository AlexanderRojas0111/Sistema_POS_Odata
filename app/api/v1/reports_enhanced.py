"""
API de Reportes Mejorada - Sistema POS Sabrositas
=================================================
Módulo robusto con exportación a Excel y visualización clara de datos
"""

from flask import Blueprint, request, jsonify, send_file
from datetime import datetime, timedelta
from sqlalchemy import func, and_, desc, text
import logging
import traceback
import io
import tempfile
import os

# Importar openpyxl para Excel
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logging.warning("openpyxl no disponible. Exportación a Excel deshabilitada.")

# Importar reportlab para PDF
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False
    logging.warning("reportlab no disponible. Exportación a PDF deshabilitada.")

from app import db
from app.models.sale import Sale, SaleItem
from app.models.product import Product
from app.models.user import User

logger = logging.getLogger(__name__)

# Blueprint mejorado
reports_enhanced_bp = Blueprint('reports_enhanced', __name__, url_prefix='/reports-enhanced')

# ===============================================
# UTILIDADES MEJORADAS
# ===============================================


def safe_execute_query(query_func, default_value=None):
    """Ejecutar consulta de forma segura con manejo de errores"""
    try:
        return query_func()
    except Exception as e:
        logger.error(f"Error en consulta: {str(e)}")
        return default_value


def format_currency(amount):
    """Formatear cantidad como moneda"""
    return float(amount) if amount is not None else 0.0


def get_date_range(start_date_str=None, end_date_str=None, default_days=30):
    """Obtener rango de fechas con valores por defecto"""
    if not start_date_str:
        start_date = datetime.now() - timedelta(days=default_days)
    else:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d')

    if not end_date_str:
        end_date = datetime.now()
    else:
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d') + timedelta(days=1)

    return start_date, end_date


def create_excel_workbook():
    """Crear workbook de Excel con estilos profesionales"""
    if not EXCEL_AVAILABLE:
        raise Exception("openpyxl no está disponible")

    wb = Workbook()

    # Estilos profesionales
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal="center", vertical="center")

    return wb, {
        'header_font': header_font,
        'header_fill': header_fill,
        'border': border,
        'center_alignment': center_alignment
    }


def create_pdf_report(report_data, report_type, filename):
    """Crear reporte PDF profesional"""
    if not PDF_AVAILABLE:
        return None

    try:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1 * inch)
        story = []

        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#1F2937')
        )

        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            spaceAfter=12,
            textColor=colors.HexColor('#374151')
        )

        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=10,
            spaceAfter=6
        )

        # Título del reporte
        story.append(Paragraph(f"Reporte de {report_type.title()}", title_style))
        story.append(Paragraph(f"Sistema POS Sabrositas - {datetime.now().strftime('%d/%m/%Y %H:%M')}",
                               styles['Normal']))
        story.append(Spacer(1, 20))

        # Resumen ejecutivo
        if 'summary' in report_data:
            story.append(Paragraph("Resumen Ejecutivo", heading_style))
            for key, value in report_data['summary'].items():
                story.append(Paragraph(f"<b>{key}:</b> {value}", normal_style))
            story.append(Spacer(1, 20))

        # Métricas principales
        if 'metrics' in report_data:
            story.append(Paragraph("Métricas Principales", heading_style))
            metrics_data = []
            for key, value in report_data['metrics'].items():
                metrics_data.append([key, str(value)])

            metrics_table = Table(metrics_data, colWidths=[3 * inch, 2 * inch])
            metrics_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3B82F6')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(metrics_table)
            story.append(Spacer(1, 20))

        # Tablas de datos
        if 'tables' in report_data:
            for table_name, table_data in report_data['tables'].items():
                story.append(Paragraph(table_name, heading_style))

                if table_data and len(table_data) > 0:
                    # Crear tabla
                    table = Table(table_data, repeatRows=1)
                    table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6B7280')),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 10),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black),
                        ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ]))
                    story.append(table)
                else:
                    story.append(Paragraph("No hay datos disponibles", normal_style))

                story.append(Spacer(1, 20))

        # Pie de página
        story.append(Spacer(1, 30))
        story.append(Paragraph("Generado por Sistema POS Sabrositas",
                               ParagraphStyle('Footer', parent=styles['Normal'],
                                              fontSize=8, alignment=TA_CENTER,
                                              textColor=colors.grey)))

        doc.build(story)
        buffer.seek(0)
        return buffer

    except Exception as e:
        logger.error(f"Error creando PDF: {str(e)}")
        return None


def apply_excel_styles(ws, styles, start_row=1, end_row=None, start_col=1, end_col=None):
    """Aplicar estilos profesionales a una hoja de Excel"""
    if end_row is None:
        end_row = ws.max_row
    if end_col is None:
        end_col = ws.max_column

    # Aplicar estilos a headers
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=start_row, column=col)
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
        cell.alignment = styles['center_alignment']
        cell.border = styles['border']

    # Aplicar bordes a todas las celdas
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            ws.cell(row=row, column=col).border = styles['border']

    # Autoajustar columnas
    for col in range(start_col, end_col + 1):
        column_letter = get_column_letter(col)
        max_length = 0
        for row in range(start_row, end_row + 1):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

# ===============================================
# ENDPOINTS PRINCIPALES MEJORADOS
# ===============================================


@reports_enhanced_bp.route('/health', methods=['GET'])
def health_check():
    """Verificación de salud del módulo mejorado"""
    try:
        # Test básico de BD
        test_result = safe_execute_query(
            lambda: db.session.execute(text('SELECT COUNT(*) FROM sales')).scalar(),
            0
        )

        return jsonify({
            'success': True,
            'status': 'healthy',
            'message': 'Módulo de reportes mejorado funcionando correctamente',
            'database_connection': 'OK',
            'sales_count': test_result,
            'excel_support': EXCEL_AVAILABLE,
            'pdf_support': PDF_AVAILABLE,
            'timestamp': datetime.now().isoformat(),
            'version': '2.0.0-enhanced'
        }), 200

    except Exception as e:
        logger.error(f"Error en health_check: {str(e)}")
        return jsonify({
            'success': False,
            'status': 'unhealthy',
            'error': str(e),
            'timestamp': datetime.now().isoformat()
        }), 500


@reports_enhanced_bp.route('/sales/analytics', methods=['GET'])
def sales_analytics():
    """Análisis avanzado de ventas con visualización clara"""
    try:
        # Parámetros
        start_date_str = request.args.get('start_date')
        end_date_str = request.args.get('end_date')
        group_by = request.args.get('group_by', 'day')  # day, hour, week, month
        include_charts = request.args.get('charts', 'true').lower() == 'true'

        start_date, end_date = get_date_range(start_date_str, end_date_str)

        # Consulta principal de ventas con información del vendedor
        sales_query = db.session.query(
            Sale.id,
            Sale.created_at,
            Sale.total_amount,
            Sale.payment_method,
            Sale.status,
            User.username.label('seller')
        ).join(User, Sale.user_id == User.id).filter(
            and_(
                Sale.created_at >= start_date,
                Sale.created_at < end_date
            )
        ).order_by(desc(Sale.created_at))

        sales = safe_execute_query(lambda: sales_query.all(), [])

        # Procesar estadísticas avanzadas
        total_sales = len(sales)
        total_revenue = sum(format_currency(sale.total_amount) for sale in sales)
        average_sale = total_revenue / total_sales if total_sales > 0 else 0

        # Análisis temporal
        temporal_analysis = {}
        payment_analysis = {}
        seller_analysis = {}

        for sale in sales:
            # Análisis temporal
            if group_by == 'day':
                key = sale.created_at.strftime('%Y-%m-%d')
            elif group_by == 'hour':
                key = sale.created_at.strftime('%H:00')
            elif group_by == 'week':
                key = sale.created_at.strftime('%Y-W%U')
            else:  # month
                key = sale.created_at.strftime('%Y-%m')

            if key not in temporal_analysis:
                temporal_analysis[key] = {'sales': 0, 'revenue': 0}
            temporal_analysis[key]['sales'] += 1
            temporal_analysis[key]['revenue'] += format_currency(sale.total_amount)

            # Análisis por método de pago
            method = sale.payment_method or 'Efectivo'
            if method not in payment_analysis:
                payment_analysis[method] = {'count': 0, 'total': 0}
            payment_analysis[method]['count'] += 1
            payment_analysis[method]['total'] += format_currency(sale.total_amount)

            # Análisis por vendedor
            seller = sale.seller or 'Desconocido'
            if seller not in seller_analysis:
                seller_analysis[seller] = {'sales': 0, 'revenue': 0}
            seller_analysis[seller]['sales'] += 1
            seller_analysis[seller]['revenue'] += format_currency(sale.total_amount)

        # Calcular tendencias
        sorted_temporal = sorted(temporal_analysis.items())
        if len(sorted_temporal) >= 2:
            recent_period = sorted_temporal[-1][1]['revenue']
            previous_period = sorted_temporal[-2][1]['revenue']
            growth_rate = ((recent_period - previous_period) / previous_period * 100) if previous_period > 0 else 0
        else:
            growth_rate = 0

        # Preparar datos para gráficos
        chart_data = {}
        if include_charts:
            chart_data = {
                'temporal_chart': {
                    'labels': [item[0] for item in sorted_temporal],
                    'sales_data': [item[1]['sales'] for item in sorted_temporal],
                    'revenue_data': [item[1]['revenue'] for item in sorted_temporal]
                },
                'payment_chart': {
                    'labels': list(payment_analysis.keys()),
                    'data': [payment_analysis[method]['total'] for method in payment_analysis.keys()]
                },
                'top_sellers': sorted(
                    [{'name': seller, 'sales': data['sales'], 'revenue': data['revenue']}
                     for seller, data in seller_analysis.items()],
                    key=lambda x: x['revenue'],
                    reverse=True
                )[:5]
            }

        response_data = {
            'success': True,
            'data': {
                'summary': {
                    'total_sales': total_sales,
                    'total_revenue': total_revenue,
                    'average_sale': round(average_sale, 2),
                    'growth_rate': round(growth_rate, 2),
                    'period': {
                        'start': start_date.strftime('%Y-%m-%d'),
                        'end': (end_date - timedelta(days=1)).strftime('%Y-%m-%d'),
                        'days': (end_date - start_date).days,
                        'group_by': group_by
                    }
                },
                'analytics': {
                    'temporal_breakdown': [
                        {
                            'period': period,
                            'sales': data['sales'],
                            'revenue': round(data['revenue'], 2),
                            'percentage': round((data['revenue'] / total_revenue * 100) if total_revenue > 0 else 0, 1)
                        }
                        for period, data in sorted_temporal
                    ],
                    'payment_methods': [
                        {
                            'method': method,
                            'count': data['count'],
                            'total': round(data['total'], 2),
                            'percentage': round((data['total'] / total_revenue * 100) if total_revenue > 0 else 0, 1)
                        }
                        for method, data in payment_analysis.items()
                    ],
                    'top_sellers': sorted(
                        [{'seller': seller, 'sales': data['sales'], 'revenue': round(data['revenue'], 2)}
                         for seller, data in seller_analysis.items()],
                        key=lambda x: x['revenue'],
                        reverse=True
                    )[:10]
                },
                'charts': chart_data if include_charts else None
            },
            'message': f'Análisis de ventas generado: {total_sales} ventas, ${total_revenue:,.2f}'
        }

        return jsonify(response_data), 200

    except Exception as e:
        logger.error(f"Error en sales_analytics: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({
            'success': False,
            'error': 'Error generando análisis de ventas',
            'details': str(e)
        }), 500


@reports_enhanced_bp.route('/inventory/analytics', methods=['GET'])
def inventory_analytics():
    """Análisis avanzado de inventario con alertas inteligentes"""
    try:
        include_details = request.args.get('details', 'false').lower() == 'true'
        category_filter = request.args.get('category')

        # Query base con información completa
        query = db.session.query(Product).filter(Product.is_active.is_(True))

        # Aplicar filtros
        if category_filter:
            query = query.filter(Product.category == category_filter)

        products = safe_execute_query(lambda: query.order_by(Product.category, Product.name).all(), [])

        # Análisis avanzado
        total_products = len(products)
        total_stock_units = sum(getattr(product, 'stock', 0) for product in products)
        total_value = sum(format_currency(getattr(product, 'price', 0) * getattr(product, 'stock', 0)) for product in products)

        # Análisis por categoría
        categories = {}
        stock_alerts = {
            'low_stock': [],
            'out_of_stock': [],
            'overstock': []
        }

        for product in products:
            stock = getattr(product, 'stock', 0)
            min_stock = getattr(product, 'min_stock', 5)
            price = format_currency(getattr(product, 'price', 0))

            # Análisis por categoría
            category = product.category or 'Sin categoría'
            if category not in categories:
                categories[category] = {
                    'product_count': 0,
                    'total_stock': 0,
                    'total_value': 0,
                    'low_stock_items': 0,
                    'out_of_stock_items': 0
                }

            categories[category]['product_count'] += 1
            categories[category]['total_stock'] += stock
            categories[category]['total_value'] += price * stock

            # Clasificar estado del stock
            if stock == 0:
                stock_alerts['out_of_stock'].append({
                    'id': product.id,
                    'name': product.name,
                    'category': category,
                    'current_stock': stock,
                    'min_stock': min_stock,
                    'price': price
                })
                categories[category]['out_of_stock_items'] += 1
            elif stock <= min_stock:
                stock_alerts['low_stock'].append({
                    'id': product.id,
                    'name': product.name,
                    'category': category,
                    'current_stock': stock,
                    'min_stock': min_stock,
                    'price': price,
                    'days_remaining': stock  # Estimación simple
                })
                categories[category]['low_stock_items'] += 1

            # Detectar overstock (más de 3x el mínimo)
            if stock > (min_stock * 3):
                stock_alerts['overstock'].append({
                    'id': product.id,
                    'name': product.name,
                    'category': category,
                    'current_stock': stock,
                    'min_stock': min_stock,
                    'price': price
                })

        response_data = {
            'success': True,
            'data': {
                'summary': {
                    'total_products': total_products,
                    'total_stock_units': total_stock_units,
                    'total_inventory_value': round(total_value, 2),
                    'categories_count': len(categories),
                    'alerts_count': {
                        'low_stock': len(stock_alerts['low_stock']),
                        'out_of_stock': len(stock_alerts['out_of_stock']),
                        'overstock': len(stock_alerts['overstock'])
                    }
                },
                'categories_analysis': [
                    {
                        'category': cat,
                        'product_count': data['product_count'],
                        'total_stock': data['total_stock'],
                        'total_value': round(data['total_value'], 2),
                        'low_stock_items': data['low_stock_items'],
                        'out_of_stock_items': data['out_of_stock_items'],
                        'percentage_of_total': round((data['total_value'] / total_value * 100) if total_value > 0 else 0, 1)
                    }
                    for cat, data in categories.items()
                ],
                'stock_alerts': stock_alerts,
                'recommendations': [
                    {
                        'type': 'reorder',
                        'priority': 'high',
                        'message': f"Reabastecer {len(stock_alerts['out_of_stock'])} productos sin stock"
                    },
                    {
                        'type': 'monitor',
                        'priority': 'medium',
                        'message': f"Monitorear {len(stock_alerts['low_stock'])} productos con stock bajo"
                    }
                ] if stock_alerts['out_of_stock'] or stock_alerts['low_stock'] else []
            },
            'message': f'Análisis de inventario: {total_products} productos, valor ${total_value:,.2f}'
        }

        # Incluir detalles si se solicita
        if include_details:
            products_details = []
            for product in products[:100]:  # Limitar para rendimiento
                stock = getattr(product, 'stock', 0)
                min_stock = getattr(product, 'min_stock', 5)
                price = format_currency(getattr(product, 'price', 0))

                # Determinar estado del stock
                if stock == 0:
                    status = 'out_of_stock'
                    urgency = 'high'
                elif stock <= min_stock:
                    status = 'low_stock'
                    urgency = 'medium'
                elif stock > (min_stock * 3):
                    status = 'overstock'
                    urgency = 'low'
                else:
                    status = 'normal'
                    urgency = 'low'

                products_details.append({
                    'id': product.id,
                    'name': product.name,
                    'sku': getattr(product, 'sku', ''),
                    'category': product.category or 'Sin categoría',
                    'price': price,
                    'current_stock': stock,
                    'min_stock': min_stock,
                    'stock_value': round(price * stock, 2),
                    'status': status,
                    'urgency': urgency,
                    'needs_reorder': stock <= min_stock
                })

            response_data['data']['products_details'] = products_details

        return jsonify(response_data), 200

    except Exception as e:
        logger.error(f"Error en inventory_analytics: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({
            'success': False,
            'error': 'Error generando análisis de inventario',
            'details': str(e)
        }), 500


@reports_enhanced_bp.route('/export/sales/excel', methods=['GET'])
def export_sales_excel():
    """Exportar ventas a Excel con formato profesional"""
    try:
        if not EXCEL_AVAILABLE:
            return jsonify({
                'success': False,
                'error': 'Exportación a Excel no disponible. Instale openpyxl.'
            }), 400

        start_date_str = request.args.get('start_date')
        end_date_str = request.args.get('end_date')

        start_date, end_date = get_date_range(start_date_str, end_date_str)

        # Obtener datos de ventas
        sales_data = safe_execute_query(
            lambda: db.session.query(
                Sale.id,
                Sale.created_at,
                Sale.total_amount,
                Sale.payment_method,
                Sale.status,
                User.username.label('seller')
            ).join(User, Sale.user_id == User.id).filter(
                and_(
                    Sale.created_at >= start_date,
                    Sale.created_at < end_date
                )
            ).order_by(Sale.created_at).all(),
            []
        )

        # Crear workbook
        wb, styles = create_excel_workbook()

        # Hoja 1: Resumen
        ws_summary = wb.active
        ws_summary.title = "Resumen de Ventas"

        # Datos del resumen
        total_sales = len(sales_data)
        total_revenue = sum(format_currency(sale.total_amount) for sale in sales_data)
        average_sale = total_revenue / total_sales if total_sales > 0 else 0

        # Escribir resumen
        ws_summary['A1'] = 'REPORTE DE VENTAS - SISTEMA POS SABROSITAS'
        ws_summary['A1'].font = Font(bold=True, size=16)
        ws_summary.merge_cells('A1:D1')

        ws_summary['A3'] = 'Período:'
        ws_summary['B3'] = f'{start_date.strftime("%Y-%m-%d")} a {(end_date - timedelta(days=1)).strftime("%Y-%m-%d")}'
        ws_summary['A4'] = 'Total de Ventas:'
        ws_summary['B4'] = total_sales
        ws_summary['A5'] = 'Ingresos Totales:'
        ws_summary['B5'] = f'${total_revenue:,.2f}'
        ws_summary['A6'] = 'Venta Promedio:'
        ws_summary['B6'] = f'${average_sale:,.2f}'
        ws_summary['A7'] = 'Generado:'
        ws_summary['B7'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # Hoja 2: Detalle de Ventas
        ws_details = wb.create_sheet("Detalle de Ventas")

        # Headers
        headers = ['ID Venta', 'Fecha', 'Hora', 'Total', 'Método de Pago', 'Estado', 'Vendedor']
        for col, header in enumerate(headers, 1):
            ws_details.cell(row=1, column=col, value=header)

        # Datos
        for row, sale in enumerate(sales_data, 2):
            ws_details.cell(row=row, column=1, value=sale.id)
            ws_details.cell(row=row, column=2, value=sale.created_at.strftime('%Y-%m-%d'))
            ws_details.cell(row=row, column=3, value=sale.created_at.strftime('%H:%M:%S'))
            ws_details.cell(row=row, column=4, value=format_currency(sale.total_amount))
            ws_details.cell(row=row, column=5, value=sale.payment_method or 'Efectivo')
            ws_details.cell(row=row, column=6, value=sale.status)
            ws_details.cell(row=row, column=7, value=sale.seller)

        # Aplicar estilos
        apply_excel_styles(ws_details, styles)

        # Hoja 3: Análisis por Método de Pago
        payment_analysis = {}
        for sale in sales_data:
            method = sale.payment_method or 'Efectivo'
            if method not in payment_analysis:
                payment_analysis[method] = {'count': 0, 'total': 0}
            payment_analysis[method]['count'] += 1
            payment_analysis[method]['total'] += format_currency(sale.total_amount)

        ws_payments = wb.create_sheet("Análisis por Pago")
        ws_payments['A1'] = 'Método de Pago'
        ws_payments['B1'] = 'Cantidad'
        ws_payments['C1'] = 'Total'
        ws_payments['D1'] = 'Porcentaje'

        row = 2
        for method, data in payment_analysis.items():
            ws_payments.cell(row=row, column=1, value=method)
            ws_payments.cell(row=row, column=2, value=data['count'])
            ws_payments.cell(row=row, column=3, value=data['total'])
            ws_payments.cell(row=row, column=4, value=f"{(data['total'] / total_revenue * 100):.1f}%")
            row += 1

        apply_excel_styles(ws_payments, styles)

        # Guardar archivo temporal
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            wb.save(tmp_file.name)

            # Leer archivo y enviar
            with open(tmp_file.name, 'rb') as f:
                file_data = f.read()

            # Limpiar archivo temporal
            os.unlink(tmp_file.name)

        # Crear respuesta
        filename = f'ventas_{start_date.strftime("%Y%m%d")}_to_{(end_date - timedelta(days=1)).strftime("%Y%m%d")}.xlsx'

        return jsonify({
            'success': True,
            'data': {
                'filename': filename,
                'file_size': len(file_data),
                'records_count': len(sales_data),
                'sheets': ['Resumen de Ventas', 'Detalle de Ventas', 'Análisis por Pago'],
                'generated_at': datetime.now().isoformat()
            },
            'message': f'Archivo Excel generado: {len(sales_data)} registros'
        }), 200

    except Exception as e:
        logger.error(f"Error en export_sales_excel: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({
            'success': False,
            'error': 'Error exportando a Excel',
            'details': str(e)
        }), 500


@reports_enhanced_bp.route('/export/inventory/excel', methods=['GET'])
def export_inventory_excel():
    """Exportar inventario a Excel con análisis completo"""
    try:
        if not EXCEL_AVAILABLE:
            return jsonify({
                'success': False,
                'error': 'Exportación a Excel no disponible. Instale openpyxl.'
            }), 400

        category_filter = request.args.get('category')

        # Query base
        query = db.session.query(Product).filter(Product.is_active.is_(True))
        if category_filter:
            query = query.filter(Product.category == category_filter)

        products = safe_execute_query(lambda: query.order_by(Product.category, Product.name).all(), [])

        # Crear workbook
        wb, styles = create_excel_workbook()

        # Hoja 1: Resumen
        ws_summary = wb.active
        ws_summary.title = "Resumen de Inventario"

        # Calcular estadísticas
        total_products = len(products)
        total_stock = sum(getattr(p, 'stock', 0) for p in products)
        total_value = sum(format_currency(getattr(p, 'price', 0) * getattr(p, 'stock', 0)) for p in products)
        low_stock_count = sum(1 for p in products if getattr(p, 'stock', 0) <= getattr(p, 'min_stock', 5))
        out_of_stock_count = sum(1 for p in products if getattr(p, 'stock', 0) == 0)

        # Escribir resumen
        ws_summary['A1'] = 'REPORTE DE INVENTARIO - SISTEMA POS SABROSITAS'
        ws_summary['A1'].font = Font(bold=True, size=16)
        ws_summary.merge_cells('A1:D1')

        ws_summary['A3'] = 'Total de Productos:'
        ws_summary['B3'] = total_products
        ws_summary['A4'] = 'Unidades en Stock:'
        ws_summary['B4'] = total_stock
        ws_summary['A5'] = 'Valor Total del Inventario:'
        ws_summary['B5'] = f'${total_value:,.2f}'
        ws_summary['A6'] = 'Productos con Stock Bajo:'
        ws_summary['B6'] = low_stock_count
        ws_summary['A7'] = 'Productos Sin Stock:'
        ws_summary['B7'] = out_of_stock_count
        ws_summary['A8'] = 'Generado:'
        ws_summary['B8'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # Hoja 2: Detalle de Productos
        ws_details = wb.create_sheet("Detalle de Productos")

        # Headers
        headers = ['ID', 'Nombre', 'SKU', 'Categoría', 'Precio', 'Stock Actual', 'Stock Mínimo', 'Valor en Stock', 'Estado']
        for col, header in enumerate(headers, 1):
            ws_details.cell(row=1, column=col, value=header)

        # Datos
        for row, product in enumerate(products, 2):
            stock = getattr(product, 'stock', 0)
            min_stock = getattr(product, 'min_stock', 5)
            price = format_currency(getattr(product, 'price', 0))

            # Determinar estado
            if stock == 0:
                status = 'Sin Stock'
            elif stock <= min_stock:
                status = 'Stock Bajo'
            else:
                status = 'Normal'

            ws_details.cell(row=row, column=1, value=product.id)
            ws_details.cell(row=row, column=2, value=product.name)
            ws_details.cell(row=row, column=3, value=getattr(product, 'sku', ''))
            ws_details.cell(row=row, column=4, value=product.category or 'Sin categoría')
            ws_details.cell(row=row, column=5, value=price)
            ws_details.cell(row=row, column=6, value=stock)
            ws_details.cell(row=row, column=7, value=min_stock)
            ws_details.cell(row=row, column=8, value=price * stock)
            ws_details.cell(row=row, column=9, value=status)

        # Aplicar estilos
        apply_excel_styles(ws_details, styles)

        # Hoja 3: Análisis por Categoría
        categories = {}
        for product in products:
            category = product.category or 'Sin categoría'
            if category not in categories:
                categories[category] = {'count': 0, 'stock': 0, 'value': 0}
            categories[category]['count'] += 1
            categories[category]['stock'] += getattr(product, 'stock', 0)
            categories[category]['value'] += format_currency(getattr(product, 'price', 0) * getattr(product, 'stock', 0))

        ws_categories = wb.create_sheet("Análisis por Categoría")
        ws_categories['A1'] = 'Categoría'
        ws_categories['B1'] = 'Productos'
        ws_categories['C1'] = 'Stock Total'
        ws_categories['D1'] = 'Valor Total'
        ws_categories['E1'] = '% del Total'

        row = 2
        for category, data in categories.items():
            percentage = (data['value'] / total_value * 100) if total_value > 0 else 0
            ws_categories.cell(row=row, column=1, value=category)
            ws_categories.cell(row=row, column=2, value=data['count'])
            ws_categories.cell(row=row, column=3, value=data['stock'])
            ws_categories.cell(row=row, column=4, value=data['value'])
            ws_categories.cell(row=row, column=5, value=f'{percentage:.1f}%')
            row += 1

        apply_excel_styles(ws_categories, styles)

        # Guardar archivo temporal
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            wb.save(tmp_file.name)

            with open(tmp_file.name, 'rb') as f:
                file_data = f.read()

            os.unlink(tmp_file.name)

        filename = f'inventario_{datetime.now().strftime("%Y%m%d")}.xlsx'

        return jsonify({
            'success': True,
            'data': {
                'filename': filename,
                'file_size': len(file_data),
                'products_count': len(products),
                'sheets': ['Resumen de Inventario', 'Detalle de Productos', 'Análisis por Categoría'],
                'generated_at': datetime.now().isoformat()
            },
            'message': f'Archivo Excel de inventario generado: {len(products)} productos'
        }), 200

    except Exception as e:
        logger.error(f"Error en export_inventory_excel: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({
            'success': False,
            'error': 'Error exportando inventario a Excel',
            'details': str(e)
        }), 500


@reports_enhanced_bp.route('/dashboard/comprehensive', methods=['GET'])
def comprehensive_dashboard():
    """Dashboard comprensivo con todas las métricas clave"""
    try:
        # Fechas de referencia
        now = datetime.now()
        today = now.date()
        yesterday = today - timedelta(days=1)
        week_ago = today - timedelta(days=7)
        month_ago = today - timedelta(days=30)

        # Métricas de ventas
        def get_sales_metrics(date_filter):
            return safe_execute_query(
                lambda: db.session.query(
                    func.count(Sale.id),
                    func.coalesce(func.sum(Sale.total_amount), 0),
                    func.avg(Sale.total_amount)
                ).filter(date_filter).first(),
                (0, 0, 0)
            )

        today_metrics = get_sales_metrics(func.date(Sale.created_at) == today)
        yesterday_metrics = get_sales_metrics(func.date(Sale.created_at) == yesterday)
        week_metrics = get_sales_metrics(Sale.created_at >= week_ago)
        month_metrics = get_sales_metrics(Sale.created_at >= month_ago)

        # Calcular tendencias
        def calculate_trend(current, previous):
            if previous > 0:
                return round(((current - previous) / previous) * 100, 1)
            return 0 if current == 0 else 100

        sales_trend = calculate_trend(format_currency(today_metrics[1]), format_currency(yesterday_metrics[1]))

        # Métricas de inventario
        inventory_stats = safe_execute_query(
            lambda: db.session.query(
                func.count(Product.id),
                func.sum(Product.stock),
                func.sum(Product.price * Product.stock),
                func.count(Product.id).filter(Product.stock <= func.coalesce(Product.min_stock, 5))
            ).filter(Product.is_active.is_(True)).first(),
            (0, 0, 0, 0)
        )

        # Productos más vendidos del mes
        top_products = safe_execute_query(
            lambda: db.session.query(
                Product.name,
                func.sum(SaleItem.quantity).label('total_sold'),
                func.sum(SaleItem.total_price).label('revenue')
            ).select_from(Product).join(SaleItem).join(Sale).filter(
                Sale.created_at >= month_ago
            ).group_by(Product.id, Product.name).order_by(
                desc(func.sum(SaleItem.quantity))
            ).limit(5).all(),
            []
        )

        # Análisis por método de pago (hoy)
        payment_today = safe_execute_query(
            lambda: db.session.query(
                Sale.payment_method,
                func.count(Sale.id),
                func.sum(Sale.total_amount)
            ).filter(func.date(Sale.created_at) == today).group_by(Sale.payment_method).all(),
            []
        )

        # Vendedores del día
        sellers_today = safe_execute_query(
            lambda: db.session.query(
                User.username,
                func.count(Sale.id),
                func.sum(Sale.total_amount)
            ).join(Sale, User.id == Sale.user_id).filter(
                func.date(Sale.created_at) == today
            ).group_by(User.id, User.username).order_by(
                desc(func.sum(Sale.total_amount))
            ).all(),
            []
        )

        response_data = {
            'success': True,
            'data': {
                'period_comparison': {
                    'today': {
                        'sales': today_metrics[0] or 0,
                        'revenue': round(format_currency(today_metrics[1]), 2),
                        'average': round(format_currency(today_metrics[2]), 2)
                    },
                    'yesterday': {
                        'sales': yesterday_metrics[0] or 0,
                        'revenue': round(format_currency(yesterday_metrics[1]), 2),
                        'average': round(format_currency(yesterday_metrics[2]), 2)
                    },
                    'week': {
                        'sales': week_metrics[0] or 0,
                        'revenue': round(format_currency(week_metrics[1]), 2),
                        'average': round(format_currency(week_metrics[2]), 2)
                    },
                    'month': {
                        'sales': month_metrics[0] or 0,
                        'revenue': round(format_currency(month_metrics[1]), 2),
                        'average': round(format_currency(month_metrics[2]), 2)
                    }
                },
                'trends': {
                    'sales_growth': sales_trend,
                    'performance': 'excellent' if sales_trend > 10 else 'good' if sales_trend > 0 else 'needs_attention'
                },
                'inventory_overview': {
                    'total_products': inventory_stats[0] or 0,
                    'total_stock_units': inventory_stats[1] or 0,
                    'total_value': round(format_currency(inventory_stats[2]), 2),
                    'low_stock_alerts': inventory_stats[3] or 0
                },
                'top_products_month': [
                    {
                        'name': product[0],
                        'quantity_sold': int(product[1] or 0),
                        'revenue': round(format_currency(product[2]), 2)
                    }
                    for product in top_products
                ],
                'payment_methods_today': [
                    {
                        'method': payment[0] or 'Efectivo',
                        'transactions': payment[1] or 0,
                        'total': round(format_currency(payment[2]), 2)
                    }
                    for payment in payment_today
                ],
                'top_sellers_today': [
                    {
                        'seller': seller[0],
                        'sales': seller[1] or 0,
                        'revenue': round(format_currency(seller[2]), 2)
                    }
                    for seller in sellers_today
                ],
                'alerts': [
                    {
                        'type': 'inventory',
                        'message': f'{inventory_stats[3] or 0} productos con stock bajo',
                        'priority': 'medium' if (inventory_stats[3] or 0) > 0 else 'low'
                    }
                ],
                'generated_at': now.isoformat()
            },
            'message': 'Dashboard comprensivo generado exitosamente'
        }

        return jsonify(response_data), 200

    except Exception as e:
        logger.error(f"Error en comprehensive_dashboard: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({
            'success': False,
            'error': 'Error generando dashboard comprensivo',
            'details': str(e)
        }), 500


@reports_enhanced_bp.route('/export/pdf', methods=['GET'])
def export_pdf():
    """Exportar reporte completo a PDF"""
    try:
        if not PDF_AVAILABLE:
            return jsonify({
                'success': False,
                'error': 'Exportación a PDF no disponible. Instale reportlab.'
            }), 400

        report_type = request.args.get('type', 'dashboard')

        # Obtener datos del dashboard
        dashboard_response = comprehensive_dashboard()
        if dashboard_response[1] != 200:
            return dashboard_response

        dashboard_data = dashboard_response[0].get_json()

        # Crear PDF
        filename = f'reporte_profesional_{report_type}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        pdf_buffer = create_pdf_report(dashboard_data, report_type, filename)

        if not pdf_buffer:
            return jsonify({
                'success': False,
                'error': 'Error generando PDF'
            }), 500

        return send_file(
            pdf_buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        logger.error(f"Error exportando a PDF: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({
            'success': False,
            'error': 'Error exportando a PDF',
            'details': str(e)
        }), 500
